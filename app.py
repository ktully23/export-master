# Author: Kari Tully, 2018

from flask import Flask, render_template, send_file, url_for
import mysql.connector
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from flask_bootstrap import Bootstrap
import random

app = Flask(__name__)
bootstrap = Bootstrap(app)

#identifying column locations for each selection
ID_COL = 26
FIRST_NAME_COL = 23
LAST_NAME_COL = 28
ADDRESS_COL = 3
CITY_COL = 6
STATE_COL = 44
ZIP_COL = 49
PHONE_COL = 34
EMAIL_COL = 16


#database connection
cnx = mysql.connector.connect(user='cclsJob', password='456CCLSjob', database='long_job_application', host='localhost')
curA = cnx.cursor(dictionary=True)
curB = cnx.cursor()
query = "SELECT firstname, lastname, address, city, state, zip, phone, email FROM job_application WHERE exported = 0;"


@app.route('/') 
@app.route('/index')
def index():

    return render_template('index.html')


@app.route('/download_excel')
def download_excel():

    #load the existing blank excel template/workbook 
    wb = load_workbook('template.xlsx')

    #get the applicants sheet in the workbook
    ws = wb['Applicants']
    
    # set the width of the column
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['P'].width = 30
    ws.column_dimensions['W'].width = 20
    ws.column_dimensions['Z'].width = 10
    ws.column_dimensions['AB'].width = 20
    ws.column_dimensions['AH'].width = 30

    #set new file name for output
    file_name = 'applicant-data.xlsx'

    #get first letter of first name and first letter of last name to create each applicants' initials
    get_id = 'SELECT concat(left(firstname, 1), left(lastname, 1)) AppId from job_application;'
    curB.execute(get_id)
    id = curB.fetchall()

    #take the initials that were created, add a random number, and put in the correct column in excel
    counter = 2
    for i in id:
        ws.cell(row=counter, column=ID_COL, value = i[0] + str(random.randint(1,1000)))
        counter += 1


    try:
        # check if there are rows to export 
        curB.execute(query)
        result = curB.fetchall()
        num_exported = len(result)

    #there is an error with the database, stop the application and inform the user  
    except mysql.connector.Error as err:
        return 'error {}'.format(err)
    else:
        #there are rows to export, so the application can continue 
        if num_exported > 0:
            counter = 2
            
            curA.execute(query)
            results = curA.fetchall()
            for row in results:
                ws.cell(row=counter, column=ADDRESS_COL, value = row['address'])
                ws.cell(row=counter, column=FIRST_NAME_COL, value = row['firstname'])
                ws.cell(row=counter, column=LAST_NAME_COL, value = row['lastname'])
                ws.cell(row=counter, column=CITY_COL, value = row['city'])
                ws.cell(row=counter, column=STATE_COL, value = row['state'])
                ws.cell(row=counter, column=EMAIL_COL, value = row['email'])
                ws.cell(row=counter, column=PHONE_COL, value = row['phone'])
                ws.cell(row=counter, column=ZIP_COL, value = row['zip'])
                counter += 1

            
            wb.save(file_name)
            return send_file(file_name, as_attachment=True)

        else:
            return '''
            Nothing to export!
            <br>
            <a href="/index">Go back</a>
            <br>
            '''
   



@app.route('/exported')
def exported():

    update_stmt = 'UPDATE job_application SET exported = 1 WHERE exported = 0;'
    curB.execute(update_stmt) 
    updated = curB.rowcount
    cnx.commit()

    return 'updated {num_updated} rows'.format(num_updated=updated)
 



if __name__ == '__main__':
    app.run()
